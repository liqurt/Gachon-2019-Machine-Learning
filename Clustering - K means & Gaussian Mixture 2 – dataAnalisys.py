import pandas as pd
import numpy as np
import csv
from openpyxl import load_workbook, Workbook
#"C:/Users/LG/Desktop/world-development-indicators/Indicators.xlsx""Indecators"

def makeList(n,row):
    list = []
    for i in range(0,n):
        list.append(row[i].value)
    return list

def makeList2(n,row):
    list = []
    for i in range(0,n):
        list.append(row[i].value)
    return list

def attrSort(attrlist,keyvalue,filename,saveFileName,sheetName,indecate):
    wb = Workbook()
    ws = wb.active
    ws.title = sheetName

    indecator = load_workbook(filename=filename);
    sheet = indecator[sheetName]

    first_row = list(sheet.rows)[0]
    alist = []

    for i in range(0, len(keyvalue)):
        alist.append(first_row[keyvalue[i]].value)
    for i in range(0,len(attrlist)):
        alist.append(attrlist[i])
    ws.append(alist)

    wb.close()
    wb.save(saveFileName)

    indecator2 = load_workbook(filename=saveFileName);
    ws = indecator2[sheetName]
    sorce_rows = list(sheet.rows)[1:]
    dic = dict()
    count = 0
    t1 = sorce_rows[0][keyvalue[1]].value
    for k in sorce_rows:
        key1 = k[keyvalue[0]].value
        key2 = k[keyvalue[1]].value
        key0 = (str(key1)+str(key2))
        if( key0 in dic):
            dic[key0][k[indecate[0]].value] = k[indecate[1]].value
        else:
            dic[key0] = {}
            dic[key0][alist[0]] = key1
            dic[key0][alist[1]] = key2
            dic[key0][k[indecate[0]].value] = k[indecate[1]].value
        if(t1 != k[keyvalue[1]].value):
            dickey = dic.keys()
            appvalue = []
            for i in dickey:
                for j in alist:
                    if(j in dic[i] ):
                      appvalue.append(dic[i][j])
                    else:
                      appvalue.append("nan")
                ws.append(appvalue)
                del appvalue[:]
            dic = dict()
            indecator2.close()
            indecator2.save(saveFileName)
            indecator2 = load_workbook(filename=saveFileName);
            ws = indecator2[sheetName]
            t1 = k[keyvalue[1]].value

    dickey = dic.keys()
    appvalue = []
    for i in dickey:
        for j in alist:
            if (j in dic[i]):
                appvalue.append(dic[i][j])
            else:
                appvalue.append("nan")
        ws.append(appvalue)
        del appvalue[:]
    indecator2.close()
    indecator2.save(saveFileName)



def makeExcel(attr,filename,criteriarow,saveFileName,sheetName):
 wb = Workbook()
 ws = wb.active
 ws.title = sheetName

 indecator = load_workbook(filename=filename);
 sheet = indecator[sheetName]
 max_col = sheet.max_column


 wb.close()
 wb.save(saveFileName)
 count = 0;

 indecator2 = load_workbook(filename=saveFileName);
 ws = indecator2[sheetName]
 first = True
 for n in sheet.rows:
    if(first):
        ws.append(makeList(max_col, n))
        first = False

    if(n[criteriarow].value in attr):
        ws.append(makeList(max_col,n))
        count+=1

    if(count == 2000):
        indecator2.close()
        indecator2.save(saveFileName)
        count = 0;
        indecator2 = load_workbook(filename=saveFileName);
        ws = indecator2[sheetName]

 indecator2.close()
 indecator2.save(saveFileName)



'''
makeExcel(List,"C:/Users/LG/Desktop/world-development-indicators/Indicators.xlsx",3,"C:/Users/LG/Desktop/world-development-indicators/Indicators2.xlsx","Indicators")
attrSort(List,[1,4],"C:/Users/LG/Desktop/world-development-indicators/Indicators2.xlsx","C:/Users/LG/Desktop/world-development-indicators/Indicators3.xlsx","Indicators",[3,5])
'''
def attrCsv(attrlist,keyvalue,filename,saveFileName,indecate):
    raw = open(filename, 'r')
    wraw = open(saveFileName, 'a', newline='')
    writer = csv.writer(wraw)
    cooked = csv.reader(raw)
    for c in cooked:
     first_row = c
     break
    alist = []
    for i in range(0, len(keyvalue)):
        alist.append(first_row[keyvalue[i]])
    for i in range(0,len(attrlist)):
        alist.append(attrlist[i])
    writer.writerow(alist)

    sorce_rows = cooked
    dic = dict()
    for k in sorce_rows:
     t1 = k[keyvalue[1]]
     break

    for k in sorce_rows:
        key1 = k[keyvalue[0]]
        key2 = k[keyvalue[1]]
        key0 = (str(key1)+str(key2))
        if( key0 in dic):
            dic[key0][k[indecate[0]]] = k[indecate[1]]
        else:
            dic[key0] = {}
            dic[key0][alist[0]] = key1
            dic[key0][alist[1]] = key2
            dic[key0][k[indecate[0]]] = k[indecate[1]]
        if(t1 != k[keyvalue[1]]):
            dickey = dic.keys()
            appvalue = []
            for i in dickey:
                for j in alist:
                    if(j in dic[i] ):
                      appvalue.append(dic[i][j])
                    else:
                      appvalue.append("nan")
                writer.writerow(appvalue)
                del appvalue[:]
            dic = dict()
            t1 = k[keyvalue[1]]

    dickey = dic.keys()
    appvalue = []
    for i in dickey:
        for j in alist:
            if (j in dic[i]):
                appvalue.append(dic[i][j])
            else:
                appvalue.append("nan")
        writer.writerow(appvalue)
        del appvalue[:]
# make new csv file that has criteriarow elements
def makeCsv(filename,saveFileName,criteriarow, attr):
 raw =  open(filename, 'r')
 wraw = open(saveFileName, 'a', newline='')
 writer = csv.writer(wraw)
 cooked = csv.reader(raw)

 first = True
 for n in cooked:
    if (first):
        writer.writerow(n)
        first = False

    if (n[criteriarow] in attr):
        writer.writerow(n)
# make data set's unique attributs list csv file
def makeCsvUnique(filename,saveFileName,criteriarow):
 raw =  open(filename, 'r')
 wraw = open(saveFileName, 'a', newline='')
 writer = csv.writer(wraw)
 cooked = csv.reader(raw)
 attrSet = set()
 first = True
 for n in cooked:
    if (first):
        writer.writerow(n[2:4])
        first = False

    elif (n[criteriarow] not in attrSet):
        writer.writerow(n[2:4])
        attrSet.add(n[criteriarow])

# get unique indicators set from filename's csv file
def makeCsvSet(filename,criteriarow,num):
 raw =  open(filename, 'r')
 cooked = csv.reader(raw)
 attrlist = []
 count = 0
 for n in cooked:
    if (count<num):
        count+=1;

    elif (n[criteriarow] not in attrlist):
        attrlist.append(n[criteriarow])
 return attrlist

# divide filename parameter's csv file by label parameter and save each results as new csv file
def deviedContry(filename,savefilepath,savefilename,labels,label):
    raw = open(filename,"r")
    cooked = csv.reader(raw)
    writer = []
    saveIndex = label
    for index in saveIndex:
        wraw = open(savefilepath+"/"+savefilename+str(index)+".csv","a",newline='')
        writer.append(csv.writer(wraw))
    for n in cooked:
        for w in writer:
            w.writerow(n)
        break
    for (n,l) in zip(cooked,labels):
        writer[l].writerow(n)

# add label columns and save that result as new csv file
def makeLabeling(filename,saveFileName,label):
 raw =  open(filename, 'r')
 wraw = open(saveFileName, 'a', newline='')
 writer = csv.writer(wraw)
 cooked = csv.reader(raw)
 count = 0;
 first = True
 for n in cooked:
    if (first):
        n.append('label')
        writer.writerow(n)
        first = False

    else:
        n.append(label[count])
        writer.writerow(n)
        count+=1
# delete row that has nan data form csv file and save that result as new csv file
def selectiveNanOut(filename,savefilename,indicators):
    raw = open(filename,"r")
    wraw = open(savefilename,"a", newline='')
    writer = csv.writer(wraw)
    cooked = csv.reader(raw)
    first = True
    writeState = True
    count = 0
    indicsLabel = []
    for n in cooked:
        if(first):
            writer.writerow(n)
            first = False
            for index in range(0,len(n)):
                if(n[index] in indicators):
                    indicsLabel.append(index)
        else:
            for indicator in indicsLabel:
                if(n[indicator] == 'nan'):
                    count+=1
                    writeState = False
                    break
            if(writeState):
                writer.writerow(n)
            writeState = True
    return count

# remove indicators in  indecators parameter form some csv file and and save that as new csv file
def selectiveOut(filename,savefilename,indicators):
    raw = open(filename,"r")
    wraw = open(savefilename,"a", newline='')
    writer = csv.writer(wraw)
    cooked = csv.reader(raw)
    first = True
    indicsLabel = []
    for n in cooked:
        if(first):
            nr = []
            for index in range(0,len(n)):
                if(n[index] not in indicators):
                    nr.append(n[index])
                    indicsLabel.append(index)
            writer.writerow(nr)
            first = False
        else:
            nr = []
            for index in indicsLabel:
                nr.append(n[index])
            writer.writerow(nr)

#get csv file's columns attributs
def getColumns(filename,start,end):
    raw = open(filename,"r")
    cooked = csv.reader(raw)
    for n in cooked:
     columns = n
     break
    if(end == "end"):
        return columns[start:]
    return columns[start:end]

# calculate most big cluster's density
def score(labels,label):
    dict = {}
    for key in label:
        dict[key] = 0
    for n in labels:
        dict[n]+=1
    result = dict.values()
    re = max(result)
    return re/len(labels)

# save data to csv file
def writeToCsv(data,filename):
    w = open(filename,"a",newline='')
    writor = csv.writer(w)
    for d in data:
        writor.writerow(d)

# join indicators and it's description
def writeSpecificToCsv(data,filename,source):
    r = open(source,"r")
    w = open(filename,"a",newline='')
    reader = csv.reader(r)
    writor = csv.writer(w)
    for d in reader:
      for n in data:
          if(n in d):
              writor.writerow(d)
